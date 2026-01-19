from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from django.core.files.storage import FileSystemStorage
import pandas as pd
import os
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def index(request):
    """Página inicial com formulário de upload"""
    return render(request, 'processador/index.html')


def processar_arquivo(request):
    if request.method == 'POST' and request.FILES.getlist('arquivo'):
        arquivos = request.FILES.getlist('arquivo')
        fs = FileSystemStorage()

        resultados_gerais = []

        try:
            for arquivo in arquivos:
                # Salva cada arquivo temporariamente
                filename = fs.save(arquivo.name, arquivo)
                arquivo_path = fs.path(filename)

                abas = pd.read_excel(arquivo_path, sheet_name=None)

                for nome_aba, df in abas.items():
                    # Normalizações
                    df = df.rename(columns={'DRE ': 'DRE', 'CH total ': 'CH total'})

                    if df["CH cursada"].dtype == object:
                        df["CH cursada"] = pd.to_timedelta(
                            df["CH cursada"] + ":00"
                        ).dt.total_seconds() / 3600

                    # Preenche células mescladas
                    df['DRE'] = df['DRE'].ffill()
                    df['Nome da escola'] = df['Nome da escola'].ffill()

                    # Agrupa dentro do próprio arquivo
                    soma_por_cpf = df.groupby(
                        ["CPF", "DRE", "Nome da escola", "Cursista"],
                        as_index=False
                    )["CH cursada"].sum()

                    resultados_gerais.append(soma_por_cpf)

                # Remove o arquivo temporário depois de processar
                fs.delete(filename)

            resultado_final = pd.concat(resultados_gerais, ignore_index=True)

            resultado_final = resultado_final.groupby(
                ["CPF", "DRE", "Nome da escola", "Cursista"],
                as_index=False
            )["CH cursada"].sum()

            resultado_final = resultado_final.sort_values(["CPF", "DRE"])

            # Salva o Excel final unificado com formatacao aprimorada
            output_path = fs.path('horas_por_cpf.xlsx')

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                resultado_final.to_excel(writer, index=False, sheet_name="Resultado")

                workbook = writer.book
                worksheet = writer.sheets["Resultado"]

                # Estilo do cabecalho
                header_font = Font(bold=True, color="000000")
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                thin_side = Side(border_style="thin", color="000000")
                header_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

                for col_idx, column_name in enumerate(resultado_final.columns, start=1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = header_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Ajuste de largura de colunas com base no conteudo
                for col_idx, column_name in enumerate(resultado_final.columns, start=1):
                    coluna_series = resultado_final[column_name].astype(str)
                    max_length = max([len(str(column_name))] + coluna_series.map(len).tolist())
                    adjusted_width = min(max_length + 2, 40)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                # Formatacao numerica especifica para "CH cursada", se existir
                if "CH cursada" in resultado_final.columns:
                    ch_col_idx = resultado_final.columns.get_loc("CH cursada") + 1
                    for row in range(2, len(resultado_final) + 2):
                        cell = worksheet.cell(row=row, column=ch_col_idx)
                        cell.number_format = "0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")

                # Alinhamento padrao para demais colunas conforme tipo
                for col_idx, column_name in enumerate(resultado_final.columns, start=1):
                    if column_name == "CH cursada":
                        continue

                    serie = resultado_final[column_name]
                    if pd.api.types.is_numeric_dtype(serie):
                        horizontal_align = "right"
                    else:
                        horizontal_align = "left"

                    for row in range(2, len(resultado_final) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.alignment = Alignment(horizontal=horizontal_align, vertical="center")

                # Congela a primeira linha (cabecalho)
                worksheet.freeze_panes = "A2"

                # Aplica autofiltro em toda a tabela
                last_row = len(resultado_final) + 1
                last_col = len(resultado_final.columns)
                filtro_ref = f"A1:{get_column_letter(last_col)}{last_row}"
                worksheet.auto_filter.ref = filtro_ref

            dados = resultado_final.to_dict('records')
            total_cpfs = resultado_final["CPF"].nunique()
            total_horas = resultado_final["CH cursada"].sum()

            context = {
                'sucesso': True,
                'dados': dados,
                'total_cpfs': total_cpfs,
                'total_horas': round(total_horas, 2)
            }

            return render(request, 'processador/resultado.html', context)

        except Exception as e:
            context = {'erro': str(e)}
            return render(request, 'processador/index.html', context)

    return render(request, 'processador/index.html')



def download_resultado(request):
    """Permite baixar o arquivo processado"""
    fs = FileSystemStorage()
    output_path = fs.path('horas_por_cpf.xlsx')
    
    if os.path.exists(output_path):
        return FileResponse(
            open(output_path, 'rb'),
            as_attachment=True,
            filename='horas_por_cpf.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    else:
        return HttpResponse('Arquivo não encontrado', status=404)